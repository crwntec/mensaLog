import xlrd
import openpyxl
import re
import os
from datetime import datetime
from models import *
from database import *

ALLOWED_CATEGORIE_KEYWORDS = ["Tagesgericht", "Vegetarisch", "Pizza & Pasta", "Wok"]

init_db()

def prettify_category(category):
    cat_str = str(category).strip()
    for keyword in ALLOWED_CATEGORIE_KEYWORDS:
        if keyword.lower() in cat_str.lower():
            return keyword
    return cat_str  # Return original if no keyword matches

def prettify_meal_str(str):
    stage1 = str.strip().replace("\n", " ").replace("\r", "")
    stage2 = re.sub(r'\([^ ]*\)', '', stage1)
    return stage2

def extract_week_from_filename(filename):
    """Extract week number from filename"""
    # Try pattern like "KW02.xls" or "KW 02.xls"
    match = re.search(r'KW\s*(\d+)', filename, re.IGNORECASE)
    if match:
        return int(match.group(1))
    
    # Try pattern like "DGE Plan Mensa 24.xls"
    match = re.search(r'Mensa\s+(\d+)', filename, re.IGNORECASE)
    if match:
        return int(match.group(1))
    
    # Try any standalone number
    match = re.search(r'(\d+)', filename)
    if match:
        return int(match.group(1))
    
    return None

def parse_xls(file_path):
    """Parse old Excel format (.xls)"""
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_index(0)
    year_dir = os.path.dirname(file_path)
    year = os.path.basename(year_dir) 
    year = int(year)
    
    # Try to extract week from the sheet first
    week = None
    try:
        week_cell = sheet.cell_value(2, 0)
        if isinstance(week_cell, str) and 'KW' in week_cell.upper():
            week = int(week_cell.split()[-1])
    except:
        pass
    
    # Fallback to filename if sheet parsing failed
    if week is None:
        filename = os.path.basename(file_path)
        week = extract_week_from_filename(filename)
    
    if week is None:
        raise ValueError(f"Could not determine week number from {file_path}")
    
    # Initialize with empty days dict
    data = Mealplan(year=int(year), week=int(week), days={})
    
    # Try to find where the actual data starts
    # Check first few rows to find date row
    date_row = None
    category_col = None
    
    for check_row in range(0, 5):
        for check_col in range(0, 3):
            try:
                cell_value = sheet.cell_value(check_row, check_col)
                # Look for a cell that might be a date in the next columns
                if check_col < sheet.ncols - 1:
                    next_cell = sheet.cell_value(check_row, check_col + 1)
                    if isinstance(next_cell, float):  # Excel dates are floats
                        try:
                            xlrd.xldate.xldate_as_datetime(next_cell, workbook.datemode)
                            date_row = check_row
                            category_col = check_col
                            break
                        except:
                            pass
            except:
                pass
        if date_row is not None:
            break
    
    if date_row is None:
        # Default assumption
        date_row = 2
        category_col = 0
    
    # Iterate over columns 1-5 for Monday-Friday
    for col in range(1, 6):
        try:
            date_cell = sheet.cell_value(date_row, col)
            date_obj = xlrd.xldate.xldate_as_datetime(date_cell, workbook.datemode)
            date_iso = date_obj.date().isoformat()
            
            # Extract meals into MealDict - CREATE NEW DICT FOR EACH DAY
            meals = {}
            
            # Start from row after date row, check up to 4 meal categories
            for row in range(date_row + 1, date_row + 5):
                try:
                    if row >= sheet.nrows:
                        break
                    meal_name = prettify_meal_str(str(sheet.cell_value(row, col)))
                    category = str(sheet.cell_value(row, category_col)).strip()
                    if meal_name and meal_name != "" and category:
                        meals[category] = meal_name
                except IndexError:
                    break
            
            if meals:  # Only add if we have meals
                data.days[date_iso] = DayDict(weekday=date_obj.strftime("%A"), meals=meals)
        except Exception as e:
            print(f"  Warning: Could not parse column {col}: {e}")
            continue
    
    return data

def parse_xlsx(file_path):
    """Parse new Excel format (.xlsx)"""
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    year_dir = os.path.dirname(file_path)
    year = os.path.basename(year_dir) 
    year = int(year)

    # Try to get week number from cell
    week = None
    for row in range(1, 5):
        try:
            week_cell = sheet.cell(row, 1).value
            if week_cell and isinstance(week_cell, str) and 'KW' in week_cell.upper():
                week = int(week_cell.split()[-1])
                break
        except:
            pass
    
    # Fallback to filename
    if week is None:
        filename = os.path.basename(file_path)
        week = extract_week_from_filename(filename)
    
    if week is None:
        raise ValueError(f"Could not determine week number from {file_path}")
    
    # Initialize with empty days dict
    data = Mealplan(year=int(year), week=int(week), days={})
    
    # Find the date row
    date_row = None
    for check_row in range(1, 10): # Check first 10 rows
        for check_col in range(2, 7): # Check columns B-F
            try:
                cell_value = sheet.cell(check_row, check_col).value
                if isinstance(cell_value, datetime):
                    date_row = check_row
                    break
            except:
                pass
        if date_row:
            break
    
    if date_row is None:
        date_row = 3  # Default fallback
    
    # Iterate over columns B-F (2-6) for Monday-Friday
    for col in range(2, 10):
        try:
            date_cell = sheet.cell(date_row, col).value
            
            date_obj = None
            if isinstance(date_cell, datetime):
                date_obj = date_cell
            elif isinstance(date_cell, str):
                try:
                    date_obj = datetime.strptime(date_cell.strip(), "%d.%m.%Y")
                except ValueError:
                    try:
                        date_obj = datetime.strptime(date_cell.strip(), "%d/%m/%Y")
                    except ValueError:
                        pass
            elif isinstance(date_cell, (int, float)):
                # Handle Excel serial date format
                from datetime import timedelta
                date_obj = datetime(1899, 12, 30) + timedelta(days=float(date_cell))
            
            # If no valid date found in this column header, skip column
            if not date_obj:
                continue
            
            date_iso = date_obj.date().isoformat()
            
            # Extract meals
            meals = {}
            
            for row in range(date_row + 1, sheet.max_row + 1):
                try:
                    # Column 1 (A) contains the Category
                    category_cell = sheet.cell(row, 1).value
                    if not category_cell:
                        continue
                    cat_str = str(category_cell).strip()
                    if not any(keyword.lower() in cat_str.lower() for keyword in ALLOWED_CATEGORIE_KEYWORDS):
                        continue
                    # Current Column contains the Meal
                    meal_name_cell = sheet.cell(row, col).value
                    
                    # Criteria: Both Category and Meal must exist
                    if category_cell and meal_name_cell:
                        category = str(category_cell).strip()
                        meal_name = prettify_meal_str(str(meal_name_cell))
                        
                        # Filter out common non-meal rows (e.g., nutritional info, prices) if necessary
                        # or ensures we don't capture empty strings
                        if meal_name and meal_name != "" and category != "":
                             # Avoid overwriting if duplicates exist, or just take the last one
                            pretty_category = prettify_category(category)  # Optionally map to standardized category names
                            meals[pretty_category] = meal_name
                except Exception:
                    continue
            
            if meals:
                data.days[date_iso] = DayDict(weekday=date_obj.strftime("%A"), meals=meals)
                
        except Exception as e:
            print(f"Error parsing column {col}: {e}")
            continue
    
    return data

def parse_excel(file_path):
    """
    Parse Excel file - automatically detects .xls or .xlsx format
    """
    if file_path.endswith('.xls'):
        return parse_xls(file_path)
    elif file_path.endswith('.xlsx'):
        return parse_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file format: {file_path}")

def import_historical_data():
    """
    Import all historical XLS/XLSX files from the directory.
    Only imports weeks that don't already exist in the database.
    """
    directory = "./archive"
    
    stats = {
        'total_files': 0,
        'imported': 0,
        'skipped': 0,
        'errors': 0
    }
    
    for year in range(2018, 2027):  # Through 2026
        year_dir = os.path.join(directory, str(year))
        
        if not os.path.exists(year_dir):
            print(f"Directory not found: {year_dir}")
            continue
        
        print(f"\n[{datetime.now()}] Processing year {year}...")
        
        for filename in sorted(os.listdir(year_dir)):
            if filename.endswith('.xls') or filename.endswith('.xlsx'):
                stats['total_files'] += 1
                file_path = os.path.join(year_dir, filename)
                
                try:
                    # Parse the file
                    mealplan = parse_excel(file_path)
                    
                    # Check if this week already exists in database
                    existing = fetch_mealplan(mealplan.year, mealplan.week)
                    
                    if existing is not None:
                        print(f"  ✓ Week {mealplan.week}/{mealplan.year} already exists. Skipping {filename}")
                        stats['skipped'] += 1
                    else:
                        # Only import if we have data
                        if mealplan.days:
                            create_mealplan(mealplan)
                            print(f"  + Week {mealplan.week}/{mealplan.year} imported from {filename} ({len(mealplan.days)} days)")
                            stats['imported'] += 1
                        else:
                            print(f"  ! Week {mealplan.week}/{mealplan.year} has no data. Skipping {filename}")
                            stats['errors'] += 1
                            
                except Exception as e:
                    print(f"  ✗ Error processing {filename}: {e}")
                    stats['errors'] += 1
    
    # Print summary
    print("\n" + "="*60)
    print("IMPORT SUMMARY")
    print("="*60)
    print(f"Total files found:     {stats['total_files']}")
    print(f"Successfully imported: {stats['imported']}")
    print(f"Already existed:       {stats['skipped']}")
    print(f"Errors:                {stats['errors']}")
    print("="*60)
